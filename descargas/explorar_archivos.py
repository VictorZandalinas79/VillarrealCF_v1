#!/usr/bin/env python3
"""
Script para reparar y convertir correctamente archivos MediaCoach

PROBLEMAS DETECTADOS:
1. XLSX corruptos (Rendimiento, maximaexigencia)
2. XML jerárquicos complejos (beyondstats, maximaexigenciarevisada)  
3. Parquet mal convertidos - todo en 1 columna con separador ; (postpartidoequipos, postpartidojugador)

SOLUCIONES:
1. Intentar múltiples métodos para leer XLSX corruptos
2. Parseo manual especializado para XML
3. Separar la columna única de parquet por punto y coma
"""

import pandas as pd
import os
import glob
from typing import Dict, Any, List
import xml.etree.ElementTree as ET
import warnings
warnings.filterwarnings('ignore')

def reparar_archivos_parquet_mal_convertidos(ruta_carpeta: str) -> bool:
    """
    Repara archivos parquet que tienen todo en una sola columna separada por ;
    """
    archivos_parquet = glob.glob(os.path.join(ruta_carpeta, "*.parquet"))
    archivos_reparados = 0
    
    print(f"\n🔧 REPARANDO ARCHIVOS PARQUET MAL CONVERTIDOS")
    print(f"📁 Carpeta: {os.path.basename(ruta_carpeta)}")
    print(f"📄 Archivos encontrados: {len(archivos_parquet)}")
    
    for archivo in archivos_parquet:
        try:
            # Leer archivo parquet
            df = pd.read_parquet(archivo)
            
            # Verificar si tiene el problema (1 columna con nombre largo con ;)
            if len(df.columns) == 1:
                nombre_columna = df.columns[0]
                if ';' in nombre_columna and len(nombre_columna) > 50:
                    print(f"\n  🔍 Analizando: {os.path.basename(archivo)}")
                    print(f"    ❌ Problema detectado: Todo en 1 columna")
                    print(f"    📋 Columna actual: {nombre_columna[:100]}...")
                    
                    # Separar header por punto y coma
                    columnas_header = nombre_columna.split(';')
                    print(f"    🏷️  Columnas detectadas: {len(columnas_header)}")
                    print(f"    📝 Primeras columnas: {columnas_header[:5]}")
                    
                    # Crear DataFrame corregido
                    datos_corregidos = []
                    
                    for idx, row in df.iterrows():
                        valor_celda = str(row.iloc[0])  # Primer (y único) valor de la fila
                        if pd.isna(row.iloc[0]) or valor_celda == 'None':
                            # Fila vacía - rellenar con None
                            fila_separada = [None] * len(columnas_header)
                        else:
                            # Separar los valores por punto y coma
                            fila_separada = valor_celda.split(';')
                            
                            # Ajustar longitud si es necesario
                            if len(fila_separada) < len(columnas_header):
                                fila_separada.extend([None] * (len(columnas_header) - len(fila_separada)))
                            elif len(fila_separada) > len(columnas_header):
                                fila_separada = fila_separada[:len(columnas_header)]
                        
                        datos_corregidos.append(fila_separada)
                    
                    # Crear nuevo DataFrame con columnas separadas
                    df_corregido = pd.DataFrame(datos_corregidos, columns=columnas_header)
                    
                    print(f"    ✅ DataFrame corregido: {df_corregido.shape}")
                    print(f"    📊 Columnas finales: {list(df_corregido.columns)}")
                    
                    # Guardar archivo corregido
                    nombre_corregido = archivo.replace('.parquet', '_CORREGIDO.parquet')
                    df_corregido.to_parquet(nombre_corregido, index=False)
                    print(f"    💾 Guardado como: {os.path.basename(nombre_corregido)}")
                    
                    # Mostrar muestra de datos corregidos
                    print(f"    👀 Muestra de datos corregidos:")
                    for col in df_corregido.columns[:3]:
                        valores_no_nulos = df_corregido[col].dropna().head(3).tolist()
                        if valores_no_nulos:
                            print(f"      {col}: {valores_no_nulos}")
                    
                    archivos_reparados += 1
                    
        except Exception as e:
            print(f"    ❌ Error procesando {os.path.basename(archivo)}: {e}")
    
    print(f"\n✅ Reparación completada: {archivos_reparados}/{len(archivos_parquet)} archivos")
    return archivos_reparados > 0

def intentar_leer_xlsx_corrupto(ruta_archivo: str) -> pd.DataFrame:
    """
    Intenta múltiples métodos para leer archivos XLSX corruptos
    """
    metodos_resultados = []
    
    # Método 1: pandas con diferentes engines
    for engine in ['openpyxl', 'xlrd']:
        try:
            df = pd.read_excel(ruta_archivo, engine=engine)
            metodos_resultados.append(('pandas_' + engine, df))
            print(f"    ✅ Método pandas {engine}: {df.shape}")
            break
        except Exception as e:
            print(f"    ❌ Método pandas {engine}: {str(e)[:100]}...")
    
    # Método 2: openpyxl directo
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta_archivo, data_only=True)
        for sheet_name in wb.sheetnames[:1]:  # Solo primera hoja
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(row)
            if data:
                df = pd.DataFrame(data[1:], columns=data[0])
                metodos_resultados.append(('openpyxl_direct', df))
                print(f"    ✅ Método openpyxl directo: {df.shape}")
                break
    except Exception as e:
        print(f"    ❌ Método openpyxl directo: {str(e)[:100]}...")
    
    # Método 3: xlrd con ignore_workbook_corruption
    try:
        import xlrd
        df = pd.read_excel(ruta_archivo, engine='xlrd')
        metodos_resultados.append(('xlrd_force', df))
        print(f"    ✅ Método xlrd forzado: {df.shape}")
    except:
        print(f"    ❌ Método xlrd forzado falló")
    
    # Retornar el primer método exitoso
    if metodos_resultados:
        return metodos_resultados[0][1]
    else:
        raise Exception("Todos los métodos fallaron")

def reparar_archivos_xlsx_corruptos(ruta_carpeta: str) -> bool:
    """
    Intenta reparar archivos XLSX corruptos
    """
    archivos_xlsx = glob.glob(os.path.join(ruta_carpeta, "*.xlsx"))
    archivos_reparados = 0
    
    print(f"\n🔧 REPARANDO ARCHIVOS XLSX CORRUPTOS")
    print(f"📁 Carpeta: {os.path.basename(ruta_carpeta)}")
    print(f"📄 Archivos encontrados: {len(archivos_xlsx)}")
    
    for archivo in archivos_xlsx[:3]:  # Limitamos a 3 archivos para prueba
        try:
            print(f"\n  🔍 Procesando: {os.path.basename(archivo)}")
            df = intentar_leer_xlsx_corrupto(archivo)
            
            if df is not None and not df.empty:
                # Guardar como parquet reparado
                nombre_reparado = archivo.replace('.xlsx', '_REPARADO.parquet')
                df.to_parquet(nombre_reparado, index=False)
                print(f"    💾 Reparado y guardado como: {os.path.basename(nombre_reparado)}")
                print(f"    📊 Dimensiones: {df.shape}")
                print(f"    🏷️  Columnas: {list(df.columns)[:5]}")
                archivos_reparados += 1
            
        except Exception as e:
            print(f"    ❌ No se pudo reparar: {str(e)[:100]}...")
    
    print(f"\n✅ Reparación XLSX completada: {archivos_reparados}/{len(archivos_xlsx[:3])} archivos")
    return archivos_reparados > 0

def parsear_xml_avanzado(ruta_archivo: str) -> pd.DataFrame:
    """
    Parseo avanzado de archivos XML jerárquicos
    """
    try:
        tree = ET.parse(ruta_archivo)
        root = tree.getroot()
        
        datos_extraidos = []
        
        # Método 1: Buscar todas las instancias
        for instance in root.findall('.//instance'):
            fila = {}
            
            # Extraer atributos de la instancia
            for attr, value in instance.attrib.items():
                fila[f'instance_{attr}'] = value
            
            # Extraer elementos hijos
            for child in instance:
                if child.tag in ['start', 'end', 'label', 'code', 'ID']:
                    fila[child.tag] = child.text
                
                # Si tiene atributos, también extraerlos
                for attr, value in child.attrib.items():
                    fila[f'{child.tag}_{attr}'] = value
            
            if fila:  # Solo agregar si tiene datos
                datos_extraidos.append(fila)
        
        if datos_extraidos:
            df = pd.DataFrame(datos_extraidos)
            return df
        
        # Método 2: Extraer cualquier elemento con texto
        for elemento in root.iter():
            if elemento.text and elemento.text.strip():
                fila = {
                    'elemento': elemento.tag,
                    'valor': elemento.text.strip(),
                    'padre': elemento.getparent().tag if elemento.getparent() is not None else None
                }
                # Agregar atributos
                for attr, value in elemento.attrib.items():
                    fila[f'attr_{attr}'] = value
                
                datos_extraidos.append(fila)
        
        if datos_extraidos:
            return pd.DataFrame(datos_extraidos)
        else:
            return pd.DataFrame({'archivo': [os.path.basename(ruta_archivo)], 'estado': ['XML sin datos extraíbles']})
            
    except Exception as e:
        return pd.DataFrame({'archivo': [os.path.basename(ruta_archivo)], 'error': [str(e)]})

def reparar_archivos_xml_complejos(ruta_carpeta: str) -> bool:
    """
    Repara archivos XML con estructura jerárquica compleja
    """
    archivos_xml = glob.glob(os.path.join(ruta_carpeta, "*.xml"))
    archivos_reparados = 0
    
    print(f"\n🔧 PROCESANDO ARCHIVOS XML COMPLEJOS")
    print(f"📁 Carpeta: {os.path.basename(ruta_carpeta)}")
    print(f"📄 Archivos encontrados: {len(archivos_xml)}")
    
    for archivo in archivos_xml[:3]:  # Limitamos a 3 archivos para prueba
        try:
            print(f"\n  🔍 Procesando: {os.path.basename(archivo)}")
            df = parsear_xml_avanzado(archivo)
            
            if df is not None and not df.empty and len(df) > 0:
                # Guardar como parquet
                nombre_procesado = archivo.replace('.xml', '_PROCESADO.parquet')
                df.to_parquet(nombre_procesado, index=False)
                print(f"    💾 Procesado y guardado como: {os.path.basename(nombre_procesado)}")
                print(f"    📊 Dimensiones: {df.shape}")
                print(f"    🏷️  Columnas: {list(df.columns)[:5]}")
                
                # Mostrar muestra
                print(f"    👀 Muestra de datos:")
                for col in df.columns[:3]:
                    valores_unicos = df[col].value_counts().head(2)
                    if not valores_unicos.empty:
                        print(f"      {col}: {dict(valores_unicos)}")
                
                archivos_reparados += 1
            else:
                print(f"    ⚠️  No se pudieron extraer datos útiles")
            
        except Exception as e:
            print(f"    ❌ Error procesando: {str(e)[:100]}...")
    
    print(f"\n✅ Procesamiento XML completado: {archivos_reparados}/{len(archivos_xml[:3])} archivos")
    return archivos_reparados > 0

def reparar_carpeta_completa(ruta_carpeta: str, nombre_carpeta: str) -> Dict[str, Any]:
    """
    Repara todos los archivos de una carpeta según su tipo
    """
    resultado = {
        'carpeta': nombre_carpeta,
        'reparaciones_realizadas': [],
        'archivos_creados': []
    }
    
    print(f"\n{'='*80}")
    print(f"🛠️  REPARANDO CARPETA: {nombre_carpeta}")
    print(f"📂 Ruta: {ruta_carpeta}")
    print(f"{'='*80}")
    
    # Detectar tipo de archivos y aplicar reparación correspondiente
    if any(nombre_carpeta.lower() in s for s in ['postpartido', 'equipos', 'jugador']):
        # Carpetas con parquet mal convertidos
        if reparar_archivos_parquet_mal_convertidos(ruta_carpeta):
            resultado['reparaciones_realizadas'].append('parquet_corregido')
            
    elif any(nombre_carpeta.lower() in s for s in ['rendimiento', 'maxima', 'exigencia']):
        # Carpetas con XLSX corruptos
        if reparar_archivos_xlsx_corruptos(ruta_carpeta):
            resultado['reparaciones_realizadas'].append('xlsx_reparado')
            
    elif any(nombre_carpeta.lower() in s for s in ['beyond', 'stats', 'revisada']):
        # Carpetas con XML complejos
        if reparar_archivos_xml_complejos(ruta_carpeta):
            resultado['reparaciones_realizadas'].append('xml_procesado')
    
    # Buscar archivos creados
    archivos_nuevos = []
    for patron in ['*_CORREGIDO.parquet', '*_REPARADO.parquet', '*_PROCESADO.parquet']:
        archivos_nuevos.extend(glob.glob(os.path.join(ruta_carpeta, patron)))
    
    resultado['archivos_creados'] = [os.path.basename(f) for f in archivos_nuevos]
    
    return resultado

def reparar_estructura_completa(ruta_base: str = "./VCF_Mediacoach_Data") -> Dict[str, Any]:
    """
    Repara toda la estructura de archivos MediaCoach
    """
    print(f"🚑 REPARADOR DE ARCHIVOS MEDIACOACH")
    print(f"📂 Ruta base: {ruta_base}")
    print("="*80)
    
    resultados_totales = {
        'carpetas_procesadas': 0,
        'reparaciones_exitosas': 0,
        'archivos_creados_total': 0,
        'detalles_por_carpeta': {}
    }
    
    if not os.path.exists(ruta_base):
        print(f"❌ La ruta {ruta_base} no existe")
        return resultados_totales
    
    # Buscar todas las carpetas de datos
    temporadas = [d for d in os.listdir(ruta_base) if os.path.isdir(os.path.join(ruta_base, d)) and d.startswith("Temporada")]
    
    for temporada in temporadas:
        ruta_temporada = os.path.join(ruta_base, temporada)
        competiciones = [d for d in os.listdir(ruta_temporada) if os.path.isdir(os.path.join(ruta_temporada, d))]
        
        for competicion in competiciones:
            ruta_competicion = os.path.join(ruta_temporada, competicion)
            carpetas_datos = [d for d in os.listdir(ruta_competicion) if os.path.isdir(os.path.join(ruta_competicion, d))]
            
            for carpeta in carpetas_datos:
                ruta_carpeta = os.path.join(ruta_competicion, carpeta)
                
                # Procesar la carpeta
                resultado = reparar_carpeta_completa(ruta_carpeta, carpeta)
                
                # Actualizar estadísticas
                key = f"{temporada}_{competicion}_{carpeta}"
                resultados_totales['detalles_por_carpeta'][key] = resultado
                resultados_totales['carpetas_procesadas'] += 1
                
                if resultado['reparaciones_realizadas']:
                    resultados_totales['reparaciones_exitosas'] += 1
                
                resultados_totales['archivos_creados_total'] += len(resultado['archivos_creados'])
    
    return resultados_totales

def generar_reporte_reparacion(resultados: Dict[str, Any]):
    """
    Genera un reporte detallado de las reparaciones realizadas
    """
    print(f"\n{'='*80}")
    print(f"📋 REPORTE DE REPARACIONES COMPLETADAS")
    print(f"{'='*80}")
    
    print(f"📊 Estadísticas generales:")
    print(f"   • Carpetas procesadas: {resultados['carpetas_procesadas']}")
    print(f"   • Reparaciones exitosas: {resultados['reparaciones_exitosas']}")
    print(f"   • Total archivos creados: {resultados['archivos_creados_total']}")
    print(f"   • Tasa de éxito: {(resultados['reparaciones_exitosas']/resultados['carpetas_procesadas']*100):.1f}%")
    
    print(f"\n📁 Detalles por carpeta:")
    for key, detalle in resultados['detalles_por_carpeta'].items():
        if detalle['reparaciones_realizadas']:
            print(f"   ✅ {detalle['carpeta']}")
            print(f"      └─ Reparaciones: {', '.join(detalle['reparaciones_realizadas'])}")
            print(f"      └─ Archivos creados: {len(detalle['archivos_creados'])}")
            if detalle['archivos_creados']:
                for archivo in detalle['archivos_creados'][:3]:
                    print(f"         • {archivo}")
        else:
            print(f"   ⏭️  {detalle['carpeta']}: Sin reparaciones necesarias")
    
    print(f"\n💡 SIGUIENTES PASOS:")
    print(f"   1. Revisa los archivos *_CORREGIDO.parquet, *_REPARADO.parquet, *_PROCESADO.parquet")
    print(f"   2. Si están correctos, puedes eliminar los archivos originales")
    print(f"   3. Considera consolidar los archivos reparados en un único parquet por carpeta")

def main():
    """Función principal del reparador"""
    print("🚑 REPARADOR DE ARCHIVOS MEDIACOACH")
    print("="*50)
    print("🎯 Este script repara archivos corruptos y mal convertidos")
    print("="*50)
    
    # Ejecutar reparaciones
    resultados = reparar_estructura_completa()
    
    # Generar reporte
    generar_reporte_reparacion(resultados)
    
    print(f"\n🎉 ¡REPARACIÓN COMPLETADA!")

if __name__ == "__main__":
    main()